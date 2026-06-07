#!/usr/bin/env python3
"""
Multi-run fair scenario comparison: same scenarios, same stochastic seeds for every
controller on every run; simulation dynamics stay on `quadrotor:` in the scenarios YAML.

Each *comparison run* (see `comparison_runs` in the scenarios config) may apply
`controller_model_overrides` on a copy of `configs/quadrotor_model.yaml` so MPC/CBF use a
mismatched internal model while Flightmare keeps the true `quadrotor:` block.

Plots are saved as separate figures per controller (positions, velocities, …) so curves stay readable.
By default each scenario folder also gets ``rollout_data.pkl`` (all seeds, all controllers)
for offline replotting; disable with ``--no-save-rollouts`` or ``plotting.save_rollouts: false``.

Usage:
  python scripts/compare_fair.py \\
    --scenarios_config configs/scenarious.yaml \\
    --checkpoint models/.../best_model \\
    --cbf_config configs/cbf_config.yaml \\
    --mpc_config configs/mpc_config.yaml \\
    --plot_dir fair_run_plots \\
    [--n_seeds 3] [--paper_plots] [--skip_controllers MPC] [--no_recompile_cbf]
"""
from __future__ import annotations

import argparse
import copy
import csv
import os
import pickle
import sys
import tempfile
import time
from datetime import datetime
from typing import Any, Dict, List, Optional, Tuple

import numpy as np
import yaml

_SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
_REPO_ROOT = os.path.dirname(_SCRIPT_DIR)
if _REPO_ROOT not in sys.path:
    sys.path.insert(0, _REPO_ROOT)

# Reuse rollout + env stack from compare_core.py
import scripts.compare_core as cs
from scripts._action_scaling import effective_thrust_limits, nominal_act_scaling, nominal_action_mass
from scripts.cbf_filter import CBFFilter, _load_cbf_config
from scripts.mpc_barrier_bounds import barriers_to_mpc_z_interval, mpc_pos_vectors_from_z
from scripts.custom_reward_wrapper import CustomRewardWrapper
from scripts.env_wrapper import ActionHistoryWrapper, ObservationNoiseWrapper
from scripts.mpc_controller import MPCController, _load_mpc_config
from scripts.quadrotor_model import QuadrotorModel

ALL_CONTROLLERS = cs.ALL_CONTROLLERS
CTRL_STYLE = cs.CTRL_STYLE


def _position_barriers_for_h_plot(cbf_config_path: Optional[str]) -> list:
    """Barriers and effective q as in ``CBFFilter`` (``q_eff = q - r_uav``) for h(x) plots.

    Used when the scenarios YAML has no ``position_barriers`` (e.g. all commented out) but
    RL+CBF still loads barriers from ``cbf_config``.
    """
    try:
        cfg = _load_cbf_config(cbf_config_path)
    except (OSError, KeyError, TypeError, yaml.YAMLError, ValueError):
        return []
    r_uav = float(cfg.get("r_uav", 0.0))
    rows: List[dict] = []
    for b in cfg.get("position_barriers", []) or []:
        row = dict(b)
        row["q"] = float(b["q"]) - r_uav
        rows.append(row)
    return rows


def _paper_scenario_index_groups(
    scenario_names_all: List[str],
) -> Tuple[List[int], List[int]]:
    """Indices for paper figures: (first up-to-6 non-inverted panel scenarios, inverted/upside-down)."""
    inverted = [
        i
        for i, n in enumerate(scenario_names_all)
        if "upside-down" in n.lower()
    ]
    non_inv = [i for i in range(len(scenario_names_all)) if i not in inverted]
    panel6 = non_inv[:6]
    return panel6, inverted


# Publication figures: vertical extent for position-error / z-height panels (m).
PAPER_PLOT_METERS_YMAX = 12.0

# Publication figures: consistent styles (do not reuse CTRL_STYLE colors).
PAPER_CTRL_STYLE = {
    "RL": {"color": "#2166ac", "linestyle": "-", "marker": "o", "markersize": 3},
    "RL+CBF": {"color": "#6a3d9a", "linestyle": "--", "marker": "s", "markersize": 3},
    "MPC": {"color": "#33a02c", "linestyle": "-", "marker": "^", "markersize": 3},
    "MPC+Con": {"color": "#e31a1c", "linestyle": "--", "marker": "D", "markersize": 3},
}


def _abbrev(name: str, n: int = 12) -> str:
    """Abbreviate scenario name to n chars for axis labels."""
    return name if len(name) <= n else name[: n - 1] + "…"


def _is_crashed(steps: int, max_steps: int) -> bool:
    return int(steps) < int(max_steps)


def _norm_perf(err: float, max_err: float = 10.0) -> float:
    return float(np.clip(1.0 - float(err) / max_err, 0.0, 1.0))


def _compute_thrust_asym(actions: np.ndarray) -> float:
    """Coefficient of variation across motors, averaged over timesteps."""
    if actions is None or len(actions) == 0:
        return float("nan")
    cvs = []
    for t in range(len(actions)):
        a = np.asarray(actions[t], dtype=np.float64).ravel()[:4]
        m = float(np.mean(a))
        if m < 1e-9:
            continue
        cvs.append(float(np.std(a) / m))
    return float(np.mean(cvs)) if cvs else float("nan")


def _compute_mae(positions: np.ndarray, goal_pos: np.ndarray) -> float:
    """Mean of ||p(t) − p_goal|| over the recorded trajectory (per-episode MAE)."""
    if positions is None or len(positions) == 0:
        return float("nan")
    g = np.asarray(goal_pos, dtype=np.float64).ravel()[:3]
    err = np.linalg.norm(positions - g, axis=1)
    return float(np.mean(err))


def _compute_t_conv(
    positions: np.ndarray,
    goal_pos: np.ndarray,
    sim_dt: float,
    threshold: float = 0.5,
) -> float:
    """First time (s) where ||p - p_goal|| < threshold; else NaN."""
    if positions is None or len(positions) == 0:
        return float("nan")
    g = np.asarray(goal_pos, dtype=np.float64).ravel()[:3]
    err = np.linalg.norm(positions - g, axis=1)
    idx = np.where(err < threshold)[0]
    if len(idx) == 0:
        return float("nan")
    return float(idx[0]) * float(sim_dt)


def _resolve_controllers(args: argparse.Namespace) -> List[str]:
    ctrl = list(args.controllers) if args.controllers else list(ALL_CONTROLLERS)
    skip = getattr(args, "skip_controllers", None) or []
    if skip:
        ctrl = [c for c in ctrl if c not in skip]
    if not ctrl:
        raise ValueError("No controllers selected (empty after --skip_controllers).")
    return ctrl


def _write_merged_quadrotor_model_yaml(
    base_rel_path: str,
    overrides: Dict[str, Any],
) -> str:
    """Write temp YAML with quadrotor_model = base + overrides. Returns path."""
    path = base_rel_path
    if not os.path.isabs(path):
        path = os.path.join(_REPO_ROOT, path)
    with open(path, "r") as f:
        doc = yaml.safe_load(f)
    if "quadrotor_model" not in doc:
        raise KeyError(f"Expected 'quadrotor_model' in {path}")
    qm = dict(doc["quadrotor_model"])
    for k, v in overrides.items():
        qm[k] = v
    doc["quadrotor_model"] = qm
    fd, tmp_path = tempfile.mkstemp(suffix="_quadrotor_model.yaml", text=True)
    try:
        with os.fdopen(fd, "w") as f:
            yaml.safe_dump(doc, f, default_flow_style=False, sort_keys=False)
    except Exception:
        os.unlink(tmp_path)
        raise
    return tmp_path


def _controller_model_path_for_run(
    run_cfg: Dict[str, Any],
    temp_files: List[str],
) -> Optional[str]:
    """Path to YAML for MPC/CBF QuadrotorModel, or None to use repo defaults."""
    overrides = run_cfg.get("controller_model_overrides") or {}
    if not overrides:
        return None
    base = run_cfg.get("base_quadrotor_model_path", "configs/quadrotor_model.yaml")
    p = _write_merged_quadrotor_model_yaml(base, overrides)
    temp_files.append(p)
    return p


def _legend_for_datasets(datasets: List[Tuple[str, dict]]):
    """Legend entries only for controllers present in ``datasets`` (readable single-controller plots)."""
    from matplotlib.lines import Line2D

    return [
        Line2D(
            [0],
            [0],
            color=CTRL_STYLE[lbl]["color"],
            ls=CTRL_STYLE[lbl]["ls"],
            lw=2,
            label=lbl,
        )
        for lbl, _ in datasets
    ]


def _controller_filename_segment(lbl: str) -> str:
    """Safe filename token for a controller name (e.g. ``RL+CBF`` → ``RL_CBF``)."""
    s = str(lbl).replace("+", "_")
    out = "".join(c if c.isalnum() or c == "_" else "_" for c in s)
    return out.strip("_") or "controller"


def _save_figure_positions(
    datasets: List[Tuple[str, dict]],
    sim_dt: float,
    goal_pos: Optional[np.ndarray],
    scenario_title: str,
    save_path: str,
) -> None:
    import matplotlib.pyplot as plt

    fig, axes = plt.subplots(3, 1, figsize=(10, 8), sharex=True)
    fig.suptitle(f"{scenario_title} — position vs time", fontsize=12, fontweight="bold")
    names = ["x", "y", "z"]
    for ax, i, nm in zip(axes, range(3), names):
        for lbl, d in datasets:
            s = CTRL_STYLE[lbl]
            pos = d["positions"]
            t = np.arange(len(pos)) * sim_dt
            ax.plot(t, pos[:, i], color=s["color"], lw=1.8, ls=s["ls"], label=lbl)
        if goal_pos is not None:
            ax.axhline(goal_pos[i], color="green", lw=1.0, ls=":", alpha=0.7)
        ax.set_ylabel(f"{nm} (m)")
        ax.grid(True, alpha=0.3)
    axes[-1].set_xlabel("Time (s)")
    h = _legend_for_datasets(datasets)
    ncol = max(1, min(len(h), 4))
    fig.legend(handles=h, loc="upper center", ncol=ncol, fontsize=9, frameon=True, bbox_to_anchor=(0.5, 1.02))
    plt.tight_layout(rect=[0, 0, 1, 0.96])
    os.makedirs(os.path.dirname(save_path) or ".", exist_ok=True)
    plt.savefig(save_path, dpi=150, bbox_inches="tight")
    plt.close()


def _save_figure_velocities(
    datasets: List[Tuple[str, dict]],
    sim_dt: float,
    scenario_title: str,
    save_path: str,
) -> None:
    import matplotlib.pyplot as plt

    fig, axes = plt.subplots(3, 1, figsize=(10, 8), sharex=True)
    fig.suptitle(f"{scenario_title} — velocity vs time", fontsize=12, fontweight="bold")
    names = ["v_x", "v_y", "v_z"]
    for ax, i, nm in zip(axes, range(3), names):
        for lbl, d in datasets:
            s = CTRL_STYLE[lbl]
            obs = d["obs"]
            t = np.arange(len(obs)) * sim_dt
            ax.plot(t, obs[:, 7 + i], color=s["color"], lw=1.8, ls=s["ls"], label=lbl)
        ax.set_ylabel(f"{nm} (m/s)")
        ax.grid(True, alpha=0.3)
    axes[-1].set_xlabel("Time (s)")
    h = _legend_for_datasets(datasets)
    ncol = max(1, min(len(h), 4))
    fig.legend(handles=h, loc="upper center", ncol=ncol, fontsize=9, frameon=True, bbox_to_anchor=(0.5, 1.02))
    plt.tight_layout(rect=[0, 0, 1, 0.96])
    os.makedirs(os.path.dirname(save_path) or ".", exist_ok=True)
    plt.savefig(save_path, dpi=150, bbox_inches="tight")
    plt.close()


def _save_figure_speed(
    datasets: List[Tuple[str, dict]],
    sim_dt: float,
    scenario_title: str,
    save_path: str,
) -> None:
    import matplotlib.pyplot as plt

    fig, ax = plt.subplots(figsize=(10, 4))
    for lbl, d in datasets:
        s = CTRL_STYLE[lbl]
        obs = d["obs"]
        t = np.arange(len(obs)) * sim_dt
        sp = np.linalg.norm(obs[:, 7:10], axis=1)
        ax.plot(t, sp, color=s["color"], lw=1.8, ls=s["ls"], label=lbl)
    ax.set_title(f"{scenario_title} — linear speed")
    ax.set_xlabel("Time (s)")
    ax.set_ylabel("|v| (m/s)")
    ax.grid(True, alpha=0.3)
    ax.legend(handles=_legend_for_datasets(datasets), loc="best", fontsize=9)
    plt.tight_layout()
    os.makedirs(os.path.dirname(save_path) or ".", exist_ok=True)
    plt.savefig(save_path, dpi=150, bbox_inches="tight")
    plt.close()


def _save_figure_tilt(
    datasets: List[Tuple[str, dict]],
    sim_dt: float,
    scenario_title: str,
    save_path: str,
) -> None:
    import matplotlib.pyplot as plt

    fig, ax = plt.subplots(figsize=(10, 4))
    for lbl, d in datasets:
        s = CTRL_STYLE[lbl]
        obs = d["obs"]
        t = np.arange(len(obs)) * sim_dt
        tilt = cs._quat_to_tilt_deg(obs[:, 3], obs[:, 4], obs[:, 5], obs[:, 6])
        ax.plot(t, tilt, color=s["color"], lw=1.8, ls=s["ls"], label=lbl)
    ax.set_title(f"{scenario_title} — tilt angle")
    ax.set_xlabel("Time (s)")
    ax.set_ylabel("Tilt (deg)")
    ax.grid(True, alpha=0.3)
    ax.legend(handles=_legend_for_datasets(datasets), loc="best", fontsize=9)
    plt.tight_layout()
    os.makedirs(os.path.dirname(save_path) or ".", exist_ok=True)
    plt.savefig(save_path, dpi=150, bbox_inches="tight")
    plt.close()


def _save_figure_angular_velocity(
    datasets: List[Tuple[str, dict]],
    sim_dt: float,
    scenario_title: str,
    save_path: str,
) -> None:
    import matplotlib.pyplot as plt

    fig, ax = plt.subplots(figsize=(10, 4))
    for lbl, d in datasets:
        s = CTRL_STYLE[lbl]
        obs = d["obs"]
        t = np.arange(len(obs)) * sim_dt
        ax.plot(t, np.linalg.norm(obs[:, 10:13], axis=1), color=s["color"], lw=1.8, ls=s["ls"], label=lbl)
    ax.set_title(f"{scenario_title} — angular velocity magnitude")
    ax.set_xlabel("Time (s)")
    ax.set_ylabel("|omega| (rad/s)")
    ax.grid(True, alpha=0.3)
    ax.legend(handles=_legend_for_datasets(datasets), loc="best", fontsize=9)
    plt.tight_layout()
    os.makedirs(os.path.dirname(save_path) or ".", exist_ok=True)
    plt.savefig(save_path, dpi=150, bbox_inches="tight")
    plt.close()


def _save_figure_thrust_total(
    datasets: List[Tuple[str, dict]],
    sim_dt: float,
    scenario_title: str,
    save_path: str,
) -> None:
    import matplotlib.pyplot as plt

    fig, ax = plt.subplots(figsize=(10, 4))
    for lbl, d in datasets:
        s = CTRL_STYLE[lbl]
        act = d["actions"]
        if len(act) == 0:
            continue
        t = np.arange(len(act)) * sim_dt
        ax.plot(t, np.sum(act, axis=1), color=s["color"], lw=1.5, ls=s["ls"], label=lbl)
    ax.set_title(f"{scenario_title} — total thrust")
    ax.set_xlabel("Time (s)")
    ax.set_ylabel("Sum motor thrust (N)")
    ax.grid(True, alpha=0.3)
    ax.legend(handles=_legend_for_datasets(datasets), loc="best", fontsize=9)
    plt.tight_layout()
    os.makedirs(os.path.dirname(save_path) or ".", exist_ok=True)
    plt.savefig(save_path, dpi=150, bbox_inches="tight")
    plt.close()


def _save_figure_thrust_per_motor(
    datasets: List[Tuple[str, dict]],
    sim_dt: float,
    scenario_title: str,
    save_path: str,
) -> None:
    import matplotlib.pyplot as plt

    fig, axes = plt.subplots(4, 1, figsize=(10, 10), sharex=True)
    fig.suptitle(f"{scenario_title} — per-motor thrust", fontsize=12, fontweight="bold")
    for mi in range(4):
        ax = axes[mi]
        for lbl, d in datasets:
            s = CTRL_STYLE[lbl]
            act = d["actions"]
            if len(act) == 0 or act.shape[1] <= mi:
                continue
            t = np.arange(len(act)) * sim_dt
            ax.plot(t, act[:, mi], color=s["color"], lw=1.2, ls=s["ls"], alpha=0.9)
        ax.set_ylabel(f"m{mi+1} (N)")
        ax.grid(True, alpha=0.3)
    axes[-1].set_xlabel("Time (s)")
    h = _legend_for_datasets(datasets)
    ncol = max(1, min(len(h), 4))
    fig.legend(handles=h, loc="upper center", ncol=ncol, fontsize=9, frameon=True, bbox_to_anchor=(0.5, 1.02))
    plt.tight_layout(rect=[0, 0, 1, 0.96])
    os.makedirs(os.path.dirname(save_path) or ".", exist_ok=True)
    plt.savefig(save_path, dpi=150, bbox_inches="tight")
    plt.close()


def _save_figure_cumulative_reward(
    datasets: List[Tuple[str, dict]],
    sim_dt: float,
    scenario_title: str,
    save_path: str,
) -> None:
    import matplotlib.pyplot as plt

    fig, ax = plt.subplots(figsize=(10, 4))
    for lbl, d in datasets:
        s = CTRL_STYLE[lbl]
        rew = d["rewards"]
        if len(rew) == 0:
            continue
        t = np.arange(len(rew)) * sim_dt
        ax.plot(t, np.cumsum(rew), color=s["color"], lw=1.8, ls=s["ls"], label=lbl)
    ax.set_title(f"{scenario_title} — cumulative reward")
    ax.set_xlabel("Time (s)")
    ax.set_ylabel("Sum reward")
    ax.grid(True, alpha=0.3)
    ax.legend(handles=_legend_for_datasets(datasets), loc="best", fontsize=9)
    plt.tight_layout()
    os.makedirs(os.path.dirname(save_path) or ".", exist_ok=True)
    plt.savefig(save_path, dpi=150, bbox_inches="tight")
    plt.close()


def _barrier_h_trajectory(
    obs: np.ndarray,
    barrier: dict,
    goal_pos: Optional[np.ndarray] = None,
) -> np.ndarray:
    """h(p,v) = n'p + q + kv*(n'v), same as ``HOCBFBarrier.h`` / ``CBFFilter``.

    Flightmare ``obs[0:3]`` is **position error** ``goal - p``, not world ``p``
    (see ``QuadrotorEnv::getObs``). When ``goal_pos`` is set, world position is
    ``p = goal_pos - obs[0:3]``, matching ``QuadrotorModel.state_from_observation``.
    If ``goal_pos`` is None, ``obs[0:3]`` is treated as world position (legacy).
    """
    obs = np.asarray(obs, dtype=np.float64)
    if obs.size == 0:
        return np.array([], dtype=np.float64)
    n = np.asarray(barrier["n"], dtype=np.float64).ravel()[:3]
    q = float(barrier["q"])
    kv = float(barrier.get("kv", 0.0))
    if goal_pos is not None:
        g = np.asarray(goal_pos, dtype=np.float64).ravel()[:3]
        p = g.reshape(1, 3) - obs[:, 0:3]
    else:
        p = obs[:, 0:3]
    v = obs[:, 7:10]
    return (p @ n) + q + kv * (v @ n)


def _save_figure_barrier_values(
    datasets: List[Tuple[str, dict]],
    barriers: list,
    sim_dt: float,
    scenario_title: str,
    save_path: str,
    goal_pos: Optional[np.ndarray] = None,
) -> None:
    """Plot velocity-aware barrier h(x) over time (safe set h >= 0)."""
    if not barriers:
        return
    import matplotlib.pyplot as plt

    nb = len(barriers)
    fig_h = min(3.2 * nb, 22.0)
    fig, axes = plt.subplots(nb, 1, figsize=(10, fig_h), sharex=True, squeeze=False)
    axes_flat = np.asarray(axes).reshape(-1)
    for bi, bdef in enumerate(barriers):
        ax = axes_flat[bi]
        bname = str(bdef.get("name", f"barrier_{bi}"))
        for lbl, d in datasets:
            s = CTRL_STYLE[lbl]
            obs = d.get("obs")
            if obs is None or len(obs) == 0:
                continue
            h = _barrier_h_trajectory(obs, bdef, goal_pos=goal_pos)
            t = np.arange(len(h), dtype=np.float64) * float(sim_dt)
            ax.plot(t, h, color=s["color"], lw=1.8, ls=s["ls"], label=lbl)
        ax.axhline(0.0, color="k", lw=1.0, ls="--", alpha=0.55)
        ax.set_ylabel(f"h\n({bname})", fontsize=9)
        ax.grid(True, alpha=0.3)
    axes_flat[-1].set_xlabel("Time (s)")
    fig.suptitle(
        f"{scenario_title} — barrier values h(x)  (safe: h ≥ 0)",
        fontsize=12,
        fontweight="bold",
    )
    h = _legend_for_datasets(datasets)
    ncol = max(1, min(len(h), 4))
    fig.legend(
        handles=h,
        loc="upper center",
        ncol=ncol,
        fontsize=9,
        frameon=True,
        bbox_to_anchor=(0.5, 1.02),
    )
    plt.tight_layout(rect=[0, 0, 1, 0.96])
    os.makedirs(os.path.dirname(save_path) or ".", exist_ok=True)
    plt.savefig(save_path, dpi=150, bbox_inches="tight")
    plt.close()


def _save_figure_position_error(
    datasets: List[Tuple[str, dict]],
    sim_dt: float,
    goal_pos: np.ndarray,
    scenario_title: str,
    save_path: str,
) -> None:
    import matplotlib.pyplot as plt

    fig, ax = plt.subplots(figsize=(10, 4))
    for lbl, d in datasets:
        s = CTRL_STYLE[lbl]
        pos = d["positions"]
        t = np.arange(len(pos)) * sim_dt
        err = np.linalg.norm(pos - goal_pos, axis=1)
        ax.plot(t, err, color=s["color"], lw=1.8, ls=s["ls"], label=lbl)
    ax.set_title(f"{scenario_title} — position error vs goal")
    ax.set_xlabel("Time (s)")
    ax.set_ylabel("||p - p_goal|| (m)")
    ax.grid(True, alpha=0.3)
    ax.legend(handles=_legend_for_datasets(datasets), loc="best", fontsize=9)
    plt.tight_layout()
    os.makedirs(os.path.dirname(save_path) or ".", exist_ok=True)
    plt.savefig(save_path, dpi=150, bbox_inches="tight")
    plt.close()


def _save_figure_trajectory_2d(
    datasets: List[Tuple[str, dict]],
    barriers: list,
    goal_pos: Optional[np.ndarray],
    scenario_title: str,
    save_path: str,
) -> None:
    import matplotlib.pyplot as plt
    from matplotlib.lines import Line2D

    fig, axes = plt.subplots(1, 3, figsize=(14, 4.5))
    pairs = [(("x", "y"), "x vs y"), (("y", "z"), "y vs z"), (("z", "x"), "z vs x")]
    for ax, (dims, title) in zip(axes, pairs):
        ia = {"x": 0, "y": 1, "z": 2}[dims[0]]
        ib = {"x": 0, "y": 1, "z": 2}[dims[1]]
        for lbl, d in datasets:
            s = CTRL_STYLE[lbl]
            pos = d["positions"]
            ax.plot(pos[:, ia], pos[:, ib], color=s["color"], lw=2, ls=s["ls"])
        if goal_pos is not None:
            ax.scatter([goal_pos[ia]], [goal_pos[ib]], color="green", s=100, marker="*", zorder=5)
        cs._draw_barriers_2d(ax, barriers, dims)
        ax.set_xlabel(dims[0])
        ax.set_ylabel(dims[1])
        ax.set_title(title)
        ax.grid(True, alpha=0.3)
        ax.axis("equal")
    fig.suptitle(f"{scenario_title} — trajectory projections", fontsize=12, fontweight="bold")
    h = _legend_for_datasets(datasets)
    extra = [
        Line2D(
            [0],
            [0],
            color="green",
            marker="*",
            ls="None",
            markersize=10,
            label="Goal",
        ),
    ]
    ncol = max(1, min(len(h) + 1, 5))
    fig.legend(
        handles=h + extra,
        loc="upper center",
        ncol=ncol,
        fontsize=9,
        bbox_to_anchor=(0.5, 1.08),
    )
    plt.tight_layout(rect=[0, 0, 1, 0.92])
    os.makedirs(os.path.dirname(save_path) or ".", exist_ok=True)
    plt.savefig(save_path, dpi=150, bbox_inches="tight")
    plt.close()


def _save_figure_trajectory_3d(
    datasets: List[Tuple[str, dict]],
    goal_pos: Optional[np.ndarray],
    scenario_title: str,
    save_path: str,
) -> None:
    import matplotlib.pyplot as plt

    fig = plt.figure(figsize=(8, 6))
    ax = fig.add_subplot(111, projection="3d")
    for lbl, d in datasets:
        s = CTRL_STYLE[lbl]
        pos = d["positions"]
        ax.plot(pos[:, 0], pos[:, 1], pos[:, 2], color=s["color"], lw=2, ls=s["ls"], label=lbl)
    if goal_pos is not None:
        ax.scatter([goal_pos[0]], [goal_pos[1]], [goal_pos[2]], color="green", s=120, marker="*", zorder=5)
    ax.set_xlabel("x")
    ax.set_ylabel("y")
    ax.set_zlabel("z")
    ax.set_title(f"{scenario_title} — 3D trajectory")
    ax.legend(handles=_legend_for_datasets(datasets), loc="best", fontsize=9)
    plt.tight_layout()
    os.makedirs(os.path.dirname(save_path) or ".", exist_ok=True)
    plt.savefig(save_path, dpi=150, bbox_inches="tight")
    plt.close()


def scenario_subdir_name(scenario_name: str) -> str:
    """Filesystem-safe single segment for a scenario-named folder."""
    s = "".join(c if c.isalnum() or c in "._- " else "_" for c in scenario_name)
    out = s.strip().replace(" ", "_")
    return out or "scenario"


def _rollouts_save_enabled(args: argparse.Namespace, scfg: Dict[str, Any]) -> bool:
    """Persist per-scenario rollout pickles unless ``--no-save-rollouts`` or ``plotting.save_rollouts: false``."""
    if getattr(args, "no_save_rollouts", False):
        return False
    if scfg.get("plotting", {}).get("save_rollouts") is False:
        return False
    return True


def save_scenario_rollout_bundle(
    out_dir: str,
    *,
    comparison_run_name: str,
    scenario_name: str,
    scenario_index: int,
    sim_dt: float,
    goal_pos: np.ndarray,
    max_episode_steps: int,
    controllers: List[str],
    barriers: list,
    no_sync_stochastic_seeds: bool,
    comparison_base_seed: int,
    ep_rollouts_by_seed: List[Dict[str, dict]],
    extras: Optional[Dict[str, Any]] = None,
) -> str:
    """Write ``rollout_data.pkl`` under ``out_dir`` (same layout as nested metric folders).

    ``rollouts_by_seed`` entries match ``compare_core._run_*_scenario`` return dicts
    (numpy arrays: positions, obs, actions, rewards; MPC: solve_times, solver_statuses;
    RL+CBF: qp_failures). Time step between samples is ``sim_dt`` (controller / metrics dt).
    """
    os.makedirs(out_dir, exist_ok=True)
    seed_meta: List[Dict[str, Any]] = []
    for st in range(len(ep_rollouts_by_seed)):
        rs: Optional[int] = None
        if not no_sync_stochastic_seeds:
            rs = int(cs._scenario_rollout_seed(comparison_base_seed, scenario_index, st))
        seed_meta.append({"seed_trial": st, "rollout_seed": rs})

    bundle: Dict[str, Any] = {
        "format_version": 1,
        "comparison_run_name": comparison_run_name,
        "scenario_name": scenario_name,
        "scenario_index": int(scenario_index),
        "sim_dt": float(sim_dt),
        "goal_pos": np.asarray(goal_pos, dtype=np.float64),
        "max_episode_steps": int(max_episode_steps),
        "controllers": list(controllers),
        "barriers": copy.deepcopy(barriers),
        "no_sync_stochastic_seeds": bool(no_sync_stochastic_seeds),
        "comparison_base_seed": int(comparison_base_seed),
        "seed_meta": seed_meta,
        "rollouts_by_seed": ep_rollouts_by_seed,
    }
    if extras:
        bundle["extras"] = dict(extras)

    path = os.path.join(out_dir, "rollout_data.pkl")
    with open(path, "wb") as f:
        pickle.dump(bundle, f, protocol=pickle.HIGHEST_PROTOCOL)
    return path


def save_scenario_plots_split(
    datasets: List[Tuple[str, dict]],
    sim_dt: float,
    barriers: list,
    goal_pos: Optional[np.ndarray],
    scenario_name: str,
    out_dir: str,
    *,
    nested_layout: bool = False,
    barriers_for_h_values: Optional[list] = None,
) -> None:
    """Write one PNG per metric **per controller** under ``out_dir`` (readable curves).

    If ``nested_layout`` is True, ``out_dir`` is the per-scenario folder
    (e.g. ``.../run_name/scenario_name/``); files look like
    ``metrics_RL_positions.png``, ``metrics_MPC_Con_angular_velocity.png``, etc.

    ``barriers`` is used for 2D trajectory overlays. ``barriers_for_h_values`` (if not
    None) defines which barriers to plot for ``metrics_*_barrier_values.png``; when None,
    ``barriers`` is used (so scenario barriers match trajectory drawings). When the
    scenario omits barriers, pass barriers from ``cbf_config`` via
    ``_position_barriers_for_h_plot`` so h matches the CBF filter.
    """
    os.makedirs(out_dir, exist_ok=True)
    if nested_layout:
        path_prefix = os.path.join(out_dir, "metrics")
    else:
        safe_name = "".join(c if c.isalnum() or c in "._- " else "_" for c in scenario_name)
        path_prefix = os.path.join(out_dir, safe_name.strip().replace(" ", "_"))

    for lbl, d in datasets:
        seg = _controller_filename_segment(lbl)
        base = f"{path_prefix}_{seg}"
        one = [(lbl, d)]
        _save_figure_positions(one, sim_dt, goal_pos, scenario_name, base + "_positions.png")
        _save_figure_velocities(one, sim_dt, scenario_name, base + "_velocities.png")
        _save_figure_speed(one, sim_dt, scenario_name, base + "_speed.png")
        _save_figure_tilt(one, sim_dt, scenario_name, base + "_tilt.png")
        _save_figure_angular_velocity(one, sim_dt, scenario_name, base + "_angular_velocity.png")
        _save_figure_thrust_total(one, sim_dt, scenario_name, base + "_thrust_total.png")
        _save_figure_thrust_per_motor(one, sim_dt, scenario_name, base + "_thrust_per_motor.png")
        _save_figure_cumulative_reward(one, sim_dt, scenario_name, base + "_cumulative_reward.png")
        if goal_pos is not None:
            _save_figure_position_error(one, sim_dt, goal_pos, scenario_name, base + "_position_error.png")
        h_barriers = barriers if barriers_for_h_values is None else barriers_for_h_values
        if h_barriers:
            _save_figure_barrier_values(
                one,
                h_barriers,
                sim_dt,
                scenario_name,
                base + "_barrier_values.png",
                goal_pos=goal_pos,
            )
        _save_figure_trajectory_2d(one, barriers, goal_pos, scenario_name, base + "_trajectory_2d.png")
        _save_figure_trajectory_3d(one, goal_pos, scenario_name, base + "_trajectory_3d.png")


def _run_one_comparison_run(
    run_idx: int,
    run_cfg: Dict[str, Any],
    scfg: Dict[str, Any],
    args: argparse.Namespace,
    temp_files: List[str],
) -> Tuple[List[dict], str, Dict[str, Dict[str, dict]]]:
    """Build env/controllers, execute all scenarios; return summary rows, plot dir, ep data."""
    from scripts.cbf_filter import set_cbf_acados_slack_force_rebuild

    run_name = run_cfg.get("name", f"run_{run_idx}")
    run_desc = run_cfg.get("description", "")
    ctrl_model_path = _controller_model_path_for_run(run_cfg, temp_files)

    controllers = _resolve_controllers(args)
    needs_rl = any(c in controllers for c in ["RL", "RL+CBF"])
    needs_mpc = any(c in controllers for c in ["MPC", "MPC+Con"])
    n_seeds = max(1, int(getattr(args, "n_seeds", 1)))

    if getattr(args, "no_recompile_cbf", False):
        set_cbf_acados_slack_force_rebuild(False)
    else:
        set_cbf_acados_slack_force_rebuild(True)

    goal_pos = np.array(scfg.get("goal", [0, 0, 5]), dtype=np.float64)
    max_steps = int(scfg.get("max_episode_steps", 1000))
    sim_dt = float(scfg.get("sim_dt", 0.02))
    barriers = scfg.get("position_barriers") or []
    scenarios = scfg.get("scenarios", [])
    rl_cfg = scfg.get("rl_policy", {})
    act_hist_len = rl_cfg.get("action_history_len", 0)
    deterministic = rl_cfg.get("deterministic", True)

    comparison_base_seed = args.comparison_base_seed
    if comparison_base_seed is None:
        comparison_base_seed = int(scfg.get("comparison_base_seed", 7777))

    # Action scaling pinned to the *nominal* mass (mass overrides hit sim only).
    nominal_mass = nominal_action_mass(scfg)
    act_mean, act_std = nominal_act_scaling(scfg, dtype=np.float64)

    print(f"\n{'#'*70}")
    print(f"Comparison run {run_idx}: {run_name}")
    if run_desc:
        print(f"  {run_desc}")
    if ctrl_model_path:
        print(f"  Controller model YAML (MPC/CBF): {ctrl_model_path}")
        with open(ctrl_model_path, "r") as f:
            qdoc = yaml.safe_load(f)
        print(f"    → mass={qdoc['quadrotor_model'].get('mass')}, "
              f"motor_tau={qdoc['quadrotor_model'].get('motor_tau')}")
    else:
        print("  Controller model: default configs/quadrotor_model.yaml (no overrides)")
    sim_plant_mass = float(scfg.get("quadrotor", {}).get("mass", nominal_mass))
    sim_mass_override = scfg.get("_sim_mass_override")
    effective_sim_mass = (
        float(sim_mass_override) if sim_mass_override is not None else sim_plant_mass
    )
    print(
        f"  Action scaling (frozen): nominal_mass={nominal_mass:.4f} kg "
        f"-> act_mean={act_mean[0]:.4f} N, act_std={act_std[0]:.4f} N",
    )
    if abs(effective_sim_mass - nominal_mass) > 1e-9:
        print(
            f"  Simulation plant mass: {effective_sim_mass:.4f} kg "
            f"(Δ vs nominal = {effective_sim_mass - nominal_mass:+.4f} kg)",
        )
    print(f"  n_seeds={n_seeds}")
    print(f"{'#'*70}")

    cfg = cs._build_env_cfg(scfg)
    base_env = cs._make_env(cfg)
    env = base_env

    # Apply sim-plant mass override (set by the sim_plant runner) via the
    # Flightmare setEnvMasses API *after* construction so the C++
    # act_mean_ / act_std_ stay pinned to the nominal training mass.
    sim_mass_override = scfg.get("_sim_mass_override")
    if sim_mass_override is not None:
        impl = getattr(base_env, "_impl", None)
        if impl is None:
            raise RuntimeError(
                "_sim_mass_override set but base_env has no Flightmare _impl",
            )
        mass_vec = np.array([float(sim_mass_override)], dtype=np.float32)
        if not impl.setEnvMasses(mass_vec):
            raise RuntimeError(
                f"setEnvMasses rejected {sim_mass_override} "
                "(check mass in (0, 100) per Flightmare).",
            )
        if hasattr(impl, "reinitHoverMotor"):
            impl.reinitHoverMotor(0)

    obs_noise_cfg = scfg.get("observation_noise")
    if isinstance(obs_noise_cfg, dict) and (
        obs_noise_cfg.get("position", 0) > 0 or obs_noise_cfg.get("velocity", 0) > 0
    ):
        env = ObservationNoiseWrapper(env, obs_noise_cfg)

    cr = rl_cfg.get("custom_reward")
    if cr is not None and cr.get("enabled", False):
        env = CustomRewardWrapper(env, cr)

    if act_hist_len > 0:
        env = ActionHistoryWrapper(env, act_hist_len)

    if not args.no_sync_stochastic_seeds:
        print(
            f"  Stochastic sync: comparison_base_seed={comparison_base_seed} "
            "(same disturbance + obs-noise per scenario for all comparison runs; "
            "only seed trial differs when n_seeds>1)",
        )

    # Effective per-motor thrust limits that match the Flightmare env.step
    # interface (action clipped to [-1,1] → thrust ∈ [0, act_mean+act_std]).
    eff_t_min, eff_t_max = effective_thrust_limits(scfg)

    cbf, model_quad = None, None
    if "RL+CBF" in controllers or needs_rl:
        cbf_kwargs = {}
        if args.cbf_config is not None:
            cbf_kwargs["config_path"] = args.cbf_config
        if ctrl_model_path is not None:
            cbf_kwargs["quadrotor_model_path"] = ctrl_model_path
        cbf = CBFFilter(**cbf_kwargs)
        cbf.set_thrust_limits(eff_t_min, eff_t_max)
        model_quad = cbf.model

    if model_quad is None:
        model_quad = QuadrotorModel(config_path=ctrl_model_path)
        model_quad.set_thrust_limits(eff_t_min, eff_t_max)

    model_policy = None
    if needs_rl:
        checkpoint = args.checkpoint or rl_cfg.get("checkpoint")
        if checkpoint is None:
            raise ValueError("--checkpoint required for RL controllers")
        if rl_cfg.get("normalize_obs", True):
            vecnorm_path = os.path.join(os.path.dirname(checkpoint), "vecnormalize.pkl")
            if os.path.isfile(vecnorm_path):
                from stable_baselines3.common.vec_env import VecNormalize

                env = VecNormalize.load(vecnorm_path, env)
                env.training = False
                env.norm_reward = False
        from stable_baselines3 import PPO

        model_policy = PPO.load(checkpoint, env=env)

    mpc_free, mpc_con, model_quad_mpc = None, None, None
    if needs_mpc:
        r_uav = 0.0
        try:
            r_uav = float(_load_cbf_config(args.cbf_config).get("r_uav", 0.0))
        except (OSError, KeyError, TypeError, yaml.YAMLError, ValueError):
            r_uav = 0.0
        mpc_ref = _load_mpc_config(args.mpc_config)
        dz = (float(mpc_ref["pos_min"][2]), float(mpc_ref["pos_max"][2]))
        z_lo, z_hi = barriers_to_mpc_z_interval(barriers, r_uav, default_z=dz)
        pos_min, pos_max = mpc_pos_vectors_from_z(z_lo, z_hi)
        model_quad_mpc = QuadrotorModel(config_path=ctrl_model_path)
        model_quad_mpc.set_thrust_limits(eff_t_min, eff_t_max)

        safe_label = "".join(c if c.isalnum() or c == "_" else "_" for c in run_name)

        if "MPC" in controllers:
            mpc_free = MPCController(
                mpc_config_path=args.mpc_config,
                quad_config_path=ctrl_model_path,
                constrained=False,
                solver_label=f"free_{safe_label}",
                thrust_limits=(eff_t_min, eff_t_max),
            )
        if "MPC+Con" in controllers:
            mpc_con = MPCController(
                mpc_config_path=args.mpc_config,
                quad_config_path=ctrl_model_path,
                pos_min=pos_min,
                pos_max=pos_max,
                constrained=True,
                solver_label=f"con_z_{safe_label}",
                thrust_limits=(eff_t_min, eff_t_max),
            )

    plot_root = args.plot_dir or scfg.get("plotting", {}).get("plot_dir", "fair_run_plots")
    run_plot_dir = os.path.join(plot_root, run_name)
    if args.save_plots or _rollouts_save_enabled(args, scfg):
        os.makedirs(run_plot_dir, exist_ok=True)

    summary_rows: List[dict] = []
    ep_data_by_scenario: Dict[str, Dict[str, dict]] = {}

    for si, scenario in enumerate(scenarios):
        name = scenario.get("name", f"Scenario {si}")
        state_13d = cs._build_state_13d(scenario)

        print(f"\n{'='*70}")
        print(f"[{run_name}] Scenario {si}: {name}")
        print(f"{'='*70}")

        metrics: Dict[str, Dict[str, List[float]]] = {
            c: {"err": [], "mae": [], "rew": [], "steps": [], "wall": [], "thrust_asym": [], "t_conv": []}
            for c in controllers
        }
        ep_data_first: Dict[str, dict] = {}
        ep_data_all_seeds: List[Dict[str, dict]] = []

        for seed_trial in range(n_seeds):
            ep_data: Dict[str, dict] = {}
            rollout_seed = None
            if not args.no_sync_stochastic_seeds:
                rollout_seed = cs._scenario_rollout_seed(
                    comparison_base_seed, si, seed_trial,
                )

            for ctrl in controllers:
                print(f"  Running {ctrl} (seed trial {seed_trial + 1}/{n_seeds}) ...")
                t0 = time.perf_counter()

                if ctrl == "RL":
                    ep_data[ctrl] = cs._run_rl_scenario(
                        env,
                        base_env,
                        model_policy,
                        model_quad,
                        goal_pos,
                        act_mean,
                        act_std,
                        max_steps,
                        state_13d,
                        act_hist_len,
                        deterministic,
                        rollout_seed,
                    )
                elif ctrl == "RL+CBF":
                    ep_data[ctrl] = cs._run_cbf_scenario(
                        env,
                        base_env,
                        model_policy,
                        cbf,
                        model_quad,
                        goal_pos,
                        act_mean,
                        act_std,
                        max_steps,
                        state_13d,
                        act_hist_len,
                        deterministic,
                        rollout_seed,
                    )
                elif ctrl == "MPC":
                    ep_data[ctrl] = cs._run_mpc_scenario(
                        env,
                        base_env,
                        mpc_free,
                        model_quad_mpc,
                        goal_pos,
                        act_mean,
                        act_std,
                        max_steps,
                        state_13d,
                        act_hist_len,
                        rollout_seed,
                    )
                elif ctrl == "MPC+Con":
                    ep_data[ctrl] = cs._run_mpc_scenario(
                        env,
                        base_env,
                        mpc_con,
                        model_quad_mpc,
                        goal_pos,
                        act_mean,
                        act_std,
                        max_steps,
                        state_13d,
                        act_hist_len,
                        rollout_seed,
                    )

                wall = time.perf_counter() - t0
                d = ep_data[ctrl]
                final_err = float(np.linalg.norm(d["positions"][-1] - goal_pos))
                mae_ep = _compute_mae(d["positions"], goal_pos)
                total_rew = float(np.sum(d["rewards"]))
                thrust_asym = _compute_thrust_asym(d.get("actions", np.array([])))
                t_conv = _compute_t_conv(d["positions"], goal_pos, sim_dt)
                metrics[ctrl]["err"].append(final_err)
                metrics[ctrl]["mae"].append(mae_ep)
                metrics[ctrl]["rew"].append(total_rew)
                metrics[ctrl]["steps"].append(float(d["steps"]))
                metrics[ctrl]["wall"].append(float(wall))
                metrics[ctrl]["thrust_asym"].append(thrust_asym)
                metrics[ctrl]["t_conv"].append(t_conv)

                print(
                    f"    steps={d['steps']}, reward={total_rew:.2f}, "
                    f"final_err={final_err:.4f}m, mae={mae_ep:.4f}m, wall={wall:.1f}s",
                )

            ep_data_all_seeds.append(copy.deepcopy(ep_data))
            if seed_trial == 0:
                ep_data_first = {k: ep_data[k] for k in ep_data}

        row: Dict[str, Any] = {"run": run_name, "name": name}
        for ctrl in controllers:
            m = metrics[ctrl]
            err_arr = np.array(m["err"], dtype=np.float64)
            mae_arr = np.array(m["mae"], dtype=np.float64)
            rew_arr = np.array(m["rew"], dtype=np.float64)
            st_arr = np.array(m["steps"], dtype=np.float64)
            wall_arr = np.array(m["wall"], dtype=np.float64)
            ta_arr = np.array(m["thrust_asym"], dtype=np.float64)
            tc_arr = np.array(m["t_conv"], dtype=np.float64)

            row[f"{ctrl}_err"] = float(np.mean(err_arr))
            row[f"{ctrl}_mae"] = float(np.mean(mae_arr))
            row[f"{ctrl}_rew"] = float(np.mean(rew_arr))
            row[f"{ctrl}_steps"] = float(np.mean(st_arr))
            row[f"{ctrl}_wall"] = float(np.mean(wall_arr))
            row[f"{ctrl}_thrust_asym"] = float(np.nanmean(ta_arr))
            tc_mean = float(np.nanmean(tc_arr))
            row[f"{ctrl}_t_conv"] = tc_mean if np.isfinite(tc_mean) else float("nan")
            row[f"{ctrl}_min_steps"] = int(np.min(st_arr))
            row[f"{ctrl}_any_crash"] = bool(np.any(st_arr < max_steps))

            if n_seeds > 1:
                row[f"{ctrl}_err_std"] = float(np.std(err_arr, ddof=0))
                row[f"{ctrl}_mae_std"] = float(np.std(mae_arr, ddof=0))
                row[f"{ctrl}_rew_std"] = float(np.std(rew_arr, ddof=0))
                row[f"{ctrl}_steps_std"] = float(np.std(st_arr, ddof=0))
                row[f"{ctrl}_wall_std"] = float(np.std(wall_arr, ddof=0))
                row[f"{ctrl}_thrust_asym_std"] = float(np.nanstd(ta_arr))
                tcc = tc_arr[np.isfinite(tc_arr)]
                row[f"{ctrl}_t_conv_std"] = (
                    float(np.std(tcc, ddof=0)) if len(tcc) > 1 else 0.0
                )

        summary_rows.append(row)
        ep_data_by_scenario[name] = ep_data_first

        if _rollouts_save_enabled(args, scfg):
            scenario_dir = os.path.join(run_plot_dir, scenario_subdir_name(name))
            os.makedirs(scenario_dir, exist_ok=True)
            extras: Dict[str, Any] = {}
            if hasattr(args, "physics_hz"):
                extras["physics_hz"] = float(getattr(args, "physics_hz"))
            if hasattr(args, "rl_hz"):
                extras["rl_hz"] = float(getattr(args, "rl_hz"))
            if hasattr(args, "reward_aggregate"):
                extras["reward_aggregate"] = str(getattr(args, "reward_aggregate"))
            rp = save_scenario_rollout_bundle(
                scenario_dir,
                comparison_run_name=run_name,
                scenario_name=name,
                scenario_index=si,
                sim_dt=float(sim_dt),
                goal_pos=goal_pos,
                max_episode_steps=max_steps,
                controllers=controllers,
                barriers=barriers,
                no_sync_stochastic_seeds=bool(args.no_sync_stochastic_seeds),
                comparison_base_seed=int(comparison_base_seed),
                ep_rollouts_by_seed=ep_data_all_seeds,
                extras=extras or None,
            )
            print(f"    Saved rollout data: {rp}")

        if args.save_plots:
            plot_datasets = [(c, ep_data_first[c]) for c in controllers]
            nested = getattr(args, "nested_comparison_plot_layout", False)
            if nested:
                scenario_plot_dir = os.path.join(run_plot_dir, scenario_subdir_name(name))
                os.makedirs(scenario_plot_dir, exist_ok=True)
                plot_out = scenario_plot_dir
            else:
                plot_out = run_plot_dir
            barriers_h = barriers if barriers else _position_barriers_for_h_plot(
                getattr(args, "cbf_config", None)
            )
            save_scenario_plots_split(
                plot_datasets,
                sim_dt,
                barriers,
                goal_pos,
                name,
                plot_out,
                nested_layout=nested,
                barriers_for_h_values=barriers_h,
            )

    env.close()
    set_cbf_acados_slack_force_rebuild(True)
    return summary_rows, run_plot_dir, ep_data_by_scenario


def _paper_save(fig, path_no_ext: str) -> None:
    import matplotlib.pyplot as plt

    fig.savefig(path_no_ext + ".pdf")
    fig.savefig(path_no_ext + ".png")
    plt.close(fig)


def generate_paper_plots(
    all_summaries: List[dict],
    all_ep_data: Dict[str, Dict[str, Dict[str, dict]]],
    scfg: Dict[str, Any],
    args: argparse.Namespace,
) -> None:
    """Publication-quality summary figures under plot_dir/paper/."""
    import matplotlib as mpl
    import matplotlib.pyplot as plt
    from matplotlib.lines import Line2D

    mpl.rcParams.update(
        {
            "font.family": "serif",
            "font.size": 9,
            "axes.labelsize": 9,
            "axes.titlesize": 9,
            "xtick.labelsize": 8,
            "ytick.labelsize": 8,
            "legend.fontsize": 8,
            "lines.linewidth": 1.4,
            "axes.linewidth": 0.6,
            "xtick.major.width": 0.6,
            "ytick.major.width": 0.6,
            "axes.grid": True,
            "grid.linewidth": 0.4,
            "grid.alpha": 0.4,
            "figure.dpi": 300,
            "savefig.dpi": 300,
            "savefig.bbox": "tight",
            "savefig.pad_inches": 0.02,
            "pdf.fonttype": 42,
            "ps.fonttype": 42,
        }
    )

    controllers = _resolve_controllers(args)
    plot_dir = os.path.join(args.plot_dir or ".", "paper")
    os.makedirs(plot_dir, exist_ok=True)

    max_steps = int(scfg.get("max_episode_steps", 1000))
    sim_dt = float(scfg.get("sim_dt", 0.02))
    goal_pos = np.array(scfg.get("goal", [0, 0, 10]), dtype=np.float64)
    scenarios_cfg = scfg.get("scenarios", [])
    scenario_names_all = [s.get("name", f"S{i}") for i, s in enumerate(scenarios_cfg)]
    n_seeds = max(1, int(getattr(args, "n_seeds", 1)))
    panel6, inverted_idx = _paper_scenario_index_groups(scenario_names_all)
    upright_panel_names = [scenario_names_all[i] for i in panel6]

    run_order = [r["name"] for r in scfg.get("comparison_runs", [])]
    if not run_order:
        run_order = sorted({row["run"] for row in all_summaries})

    def _row(run: str, scen: str) -> Optional[dict]:
        for row in all_summaries:
            if row["run"] == run and row["name"] == scen:
                return row
        return None

    # --- Figure 1: final error bars (double-column) ---
    n_runs = len(run_order)
    if n_runs <= 2:
        nrows, ncols = 1, n_runs
    elif n_runs == 3:
        nrows, ncols = 1, 3
    elif n_runs == 4:
        nrows, ncols = 2, 2
    else:
        ncols = 2
        nrows = int(np.ceil(n_runs / ncols))

    n_scen = len(scenario_names_all)
    fig1_w = min(24.0, max(7.16, 0.36 * n_scen + 3.0))
    fig1, axes1 = plt.subplots(nrows, ncols, figsize=(fig1_w, 2.8 * nrows), squeeze=False)
    ax_flat = axes1.ravel()
    x = np.arange(n_scen)
    width = 0.18
    offsets = np.linspace(-(len(controllers) - 1) / 2, (len(controllers) - 1) / 2, len(controllers)) * width * 1.15

    for ri, run_name in enumerate(run_order):
        if ri >= len(ax_flat):
            break
        ax = ax_flat[ri]
        for ki, ctrl in enumerate(controllers):
            hs = []
            crash_m = []
            for sj in range(n_scen):
                sn = scenario_names_all[sj]
                row = _row(run_name, sn)
                err = float(row.get(f"{ctrl}_err", float("nan"))) if row else float("nan")
                hs.append(err)
                crashed = False
                if row:
                    crashed = bool(row.get(f"{ctrl}_any_crash", False))
                    if not crashed:
                        crashed = _is_crashed(
                            int(row.get(f"{ctrl}_min_steps", row.get(f"{ctrl}_steps", 0))),
                            max_steps,
                        )
                crash_m.append(crashed)
            pos = x + offsets[ki]
            ps = PAPER_CTRL_STYLE[ctrl]
            for sj, xi in enumerate(pos):
                h = hs[sj]
                crashed = crash_m[sj]
                hb = min(h, 7.5) if np.isfinite(h) else 7.5
                alpha = 0.4 if crashed else 1.0
                ax.bar(xi, hb, width, color=ps["color"], alpha=alpha, edgecolor="none")
                if np.isfinite(h) and h < 1.0:
                    ax.text(xi, hb, f"{h:.2f} m", ha="center", va="bottom", fontsize=6)
                row = _row(run_name, scenario_names_all[sj])
                tc = row.get(f"{ctrl}_t_conv", float("nan")) if row else float("nan")
                if row is not None and n_seeds > 1 and f"{ctrl}_t_conv_std" in row:
                    tcs = row[f"{ctrl}_t_conv_std"]
                    tc_lab = (
                        f"conv: {tc:.2f}±{tcs:.2f} s"
                        if np.isfinite(tc) and np.isfinite(tcs)
                        else ("no conv" if not np.isfinite(tc) else f"conv: {tc:.2f} s")
                    )
                else:
                    tc_lab = f"conv: {tc:.2f} s" if np.isfinite(tc) else "no conv"
                ax.text(xi, -0.25, tc_lab, ha="center", va="top", fontsize=5)
                if crashed:
                    ax.plot(xi, hb + 0.2, marker="x", color="red", markersize=5, linestyle="None")
        ax.axhline(2.0, color="0.45", ls="--", lw=0.8)
        ax.text(
            0.02,
            2.05,
            "safe-set boundary",
            transform=ax.get_yaxis_transform(),
            fontsize=7,
            color="0.35",
        )
        ax.set_title(_abbrev(run_name, 20))
        ax.set_ylim(-0.9, 7.5)
        ax.set_xticks(x)
        ax.set_xticklabels(
            [_abbrev(s, 11) for s in scenario_names_all],
            rotation=58,
            ha="right",
            fontsize=6,
        )
        if ri % ncols == 0:
            ax.set_ylabel("final ||p − p*|| (m)")

    for j in range(len(run_order), len(ax_flat)):
        ax_flat[j].set_visible(False)

    handles = [
        Line2D([0], [0], color=PAPER_CTRL_STYLE[c]["color"], lw=4, label=c) for c in controllers
    ]
    fig1.legend(handles=handles, loc="lower center", ncol=len(controllers), bbox_to_anchor=(0.5, -0.02))
    fig1.tight_layout(rect=[0, 0.06, 1, 1])
    _paper_save(fig1, os.path.join(plot_dir, "fig_final_error_bar"))

    # --- Figure 2 & 3: per comparison run, first non-inverted scenarios (2×3 grid, up to 6) ---
    n_panel = min(6, len(upright_panel_names))

    if n_panel > 0:
        for run_name in run_order:
            safe_rn = "".join(ch if ch.isalnum() or ch in "-_" else "_" for ch in str(run_name))[:48]
            fig2, axes2 = plt.subplots(2, 3, figsize=(7.16, 5.2), squeeze=False)
            fig3, axes3 = plt.subplots(2, 3, figsize=(7.16, 5.2), squeeze=False)
            for pi in range(6):
                r, c = divmod(pi, 3)
                ax_e = axes2[r, c]
                ax_z = axes3[r, c]
                if pi >= n_panel:
                    ax_e.set_visible(False)
                    ax_z.set_visible(False)
                    continue
                scen = upright_panel_names[pi]
                ed_run = all_ep_data.get(run_name, {})
                t_max = sim_dt
                for ctrl in controllers:
                    ps = PAPER_CTRL_STYLE[ctrl]
                    d = ed_run.get(scen, {}).get(ctrl)
                    if d is None:
                        continue
                    pos = d["positions"]
                    t = np.arange(len(pos)) * sim_dt
                    t_max = max(t_max, float(t[-1]) if len(t) else 0.0)
                    err = np.linalg.norm(pos - goal_pos, axis=1)
                    ax_e.plot(
                        t,
                        err,
                        color=ps["color"],
                        ls=ps["linestyle"],
                        marker=ps["marker"],
                        ms=ps["markersize"],
                        markevery=max(1, len(t) // 20),
                        label=ctrl,
                    )
                    fe = float(err[-1])
                    ax_e.text(
                        t[-1],
                        fe,
                        f" {fe:.2f}",
                        fontsize=6,
                        color=ps["color"],
                        va="center",
                    )
                    z = pos[:, 2]
                    ax_z.plot(
                        t,
                        z,
                        color=ps["color"],
                        ls=ps["linestyle"],
                        marker=ps["marker"],
                        ms=ps["markersize"],
                        markevery=max(1, len(t) // 20),
                        label=ctrl,
                    )
                ax_e.axhline(2.0, color="0.5", ls="--", lw=0.7)
                ax_e.fill_between(
                    [0.0, t_max],
                    2.0,
                    PAPER_PLOT_METERS_YMAX,
                    color="0.5",
                    alpha=0.08,
                )
                ax_e.set_xlim(0.0, max(t_max, sim_dt))
                ax_e.set_ylim(0.0, PAPER_PLOT_METERS_YMAX)
                ax_e.set_title(_abbrev(scen, 20))
                ax_z.set_title(_abbrev(scen, 20))
                if pi in (0, 1):
                    ax_e.set_ylabel("||p − p*|| (m)")
                if pi in (3, 4):
                    ax_e.set_xlabel("time (s)")
                if pi in (0, 1):
                    ax_z.set_ylabel("z (m)")
                if pi in (3, 4):
                    ax_z.set_xlabel("time (s)")
                ax_z.axhline(8.0, color="red", ls="-", lw=0.9)
                ax_z.axhline(10.0, color="green", ls="--", lw=0.8)
                ax_z.fill_between(
                    [0.0, t_max],
                    8.0,
                    PAPER_PLOT_METERS_YMAX,
                    color="red",
                    alpha=0.08,
                )
                ax_z.set_xlim(0.0, max(t_max, sim_dt))
                ax_z.set_ylim(0.0, PAPER_PLOT_METERS_YMAX)
            h2 = [
                Line2D([0], [0], color=PAPER_CTRL_STYLE[c]["color"], ls=PAPER_CTRL_STYLE[c]["linestyle"], label=c)
                for c in controllers
            ]
            fig2.legend(handles=h2, loc="lower center", ncol=4, bbox_to_anchor=(0.5, -0.02))
            fig2.suptitle(f"Position error — {run_name}", y=1.02, fontsize=9)
            fig2.tight_layout(rect=[0, 0.05, 1, 0.98])
            _paper_save(fig2, os.path.join(plot_dir, f"fig_error_timeseries_{safe_rn}"))

            fig3.text(0.5, 0.01, "CBF/MPC barrier: z=8 m (red)  |  goal: z=10 m (green)", ha="center", fontsize=7)
            h3 = [
                Line2D([0], [0], color=PAPER_CTRL_STYLE[c]["color"], ls=PAPER_CTRL_STYLE[c]["linestyle"], label=c)
                for c in controllers
            ]
            fig3.legend(handles=h3, loc="lower center", ncol=4, bbox_to_anchor=(0.5, -0.02))
            fig3.suptitle(f"z height — {run_name}", y=1.02, fontsize=9)
            fig3.tight_layout(rect=[0, 0.06, 1, 0.98])
            _paper_save(fig3, os.path.join(plot_dir, f"fig_z_height_{safe_rn}"))

    # --- Figure 4: radar (mean over runs) ---
    if not upright_panel_names:
        pass
    else:
        n_spokes = len(upright_panel_names)
        angles = np.linspace(0, 2 * np.pi, n_spokes, endpoint=False).tolist()
        angles_closed = angles + angles[:1]
        fig4 = plt.figure(figsize=(3.5, 4.2))
        ax_r = fig4.add_subplot(211, polar=True)
        ax_r.set_theta_offset(np.pi / 2)
        ax_r.set_theta_direction(-1)

        for ctrl in controllers:
            vals = []
            for scen in upright_panel_names:
                perf_vals = []
                for run_name in run_order:
                    row = _row(run_name, scen)
                    if not row:
                        continue
                    err = float(row.get(f"{ctrl}_err", 10.0))
                    perf_vals.append(_norm_perf(err))
                if perf_vals:
                    vals.append(float(np.mean(perf_vals)))
                else:
                    vals.append(0.0)
            vals_closed = vals + vals[:1]
            ps = PAPER_CTRL_STYLE[ctrl]
            ax_r.plot(angles_closed, vals_closed, color=ps["color"], ls=ps["linestyle"], label=ctrl)
            ax_r.fill(angles_closed, vals_closed, color=ps["color"], alpha=0.15)

        ax_r.set_xticks(angles)
        ax_r.set_xticklabels([_abbrev(s, 9) for s in upright_panel_names], fontsize=6)
        ax_r.set_ylim(0, 1)
        ax_r.set_title("Robustness (mean across runs)", y=1.08, fontsize=9)
        ax_r.legend(loc="upper right", bbox_to_anchor=(1.35, 1.1), fontsize=7)

        ax_tbl = fig4.add_subplot(212)
        ax_tbl.axis("off")
        table_data = [["Controller", "mean ± std (norm)"]]
        for ctrl in controllers:
            cell_vals = []
            for run_name in run_order:
                for scen in upright_panel_names:
                    row = _row(run_name, scen)
                    if row:
                        err = float(row.get(f"{ctrl}_err", 10.0))
                        cell_vals.append(_norm_perf(err))
            if cell_vals:
                m, s = float(np.mean(cell_vals)), float(np.std(cell_vals, ddof=0))
                table_data.append([ctrl, f"{m:.3f} ± {s:.3f}"])
            else:
                table_data.append([ctrl, "—"])
        tbl = ax_tbl.table(
            cellText=table_data,
            loc="center",
            cellLoc="center",
            colWidths=[0.35, 0.45],
        )
        tbl.auto_set_font_size(False)
        tbl.set_fontsize(7)
        fig4.tight_layout()
        _paper_save(fig4, os.path.join(plot_dir, "fig_radar_robustness"))

    # --- Figure 5: wall time per step ---
    ms_samples: Dict[str, List[float]] = {c: [] for c in controllers}
    for row in all_summaries:
        for ctrl in controllers:
            st = row.get(f"{ctrl}_steps", max_steps)
            w = row.get(f"{ctrl}_wall", 0.0)
            if st and int(row.get(f"{ctrl}_min_steps", st)) >= max_steps:
                ms_samples[ctrl].append(1000.0 * float(w) / float(st))

    means = {c: float(np.mean(ms_samples[c])) if ms_samples[c] else float("nan") for c in controllers}
    fig5, ax5 = plt.subplots(figsize=(3.5, 2.8))
    y_pos = np.arange(len(controllers))
    for i, ctrl in enumerate(controllers):
        v = means[ctrl]
        v_plot = v if np.isfinite(v) and v > 0 else np.nan
        if np.isfinite(v_plot):
            ax5.barh(i, v_plot, color=PAPER_CTRL_STYLE[ctrl]["color"], height=0.55)
            ax5.text(v_plot * 1.08, i, f"{v:.2f} ms", va="center", fontsize=7)
        else:
            ax5.text(1e-3, i, "n/a (no full episodes)", va="center", fontsize=7)
    ax5.set_yticks(y_pos)
    ax5.set_yticklabels(controllers)
    ax5.set_xscale("log")
    ax5.set_xlabel("mean wall time per step (ms)")
    ax5.set_title("Mean compute time per control step")
    fig5.tight_layout()
    _paper_save(fig5, os.path.join(plot_dir, "fig_wall_time"))

    # --- Figure 6: mismatch heatmap ---
    mismatch_runs = [r["name"] for r in scfg.get("comparison_runs", []) if r.get("name") != "nominal"]
    n_mat_cols = len(mismatch_runs) * len(panel6)
    matrix = np.zeros((len(controllers), max(1, n_mat_cols)))
    xtick_labels: List[str] = []
    col_idx = 0
    for mr in mismatch_runs:
        for ii in panel6:
            if ii >= len(scenario_names_all):
                continue
            xtick_labels.append(_abbrev(scenario_names_all[ii], 9))
            for ci, ctrl in enumerate(controllers):
                sn = scenario_names_all[ii]
                row = _row(mr, sn)
                err = float(row.get(f"{ctrl}_err", 7.0)) if row else 7.0
                crashed = bool(row.get(f"{ctrl}_any_crash", False)) if row else False
                val = float(np.clip(err, 0.0, 7.0))
                if crashed:
                    val = 7.0
                matrix[ci, col_idx] = val
            col_idx += 1
    matrix = matrix[:, :col_idx] if col_idx > 0 else matrix
    if matrix.shape[1] > 0 and mismatch_runs:
        fig6, ax6 = plt.subplots(figsize=(7.16, 2.8))
        im = ax6.imshow(matrix, aspect="auto", cmap="RdYlGn_r", vmin=0, vmax=7)
        ax6.set_yticks(range(len(controllers)))
        ax6.set_yticklabels(controllers)
        ax6.set_xticks(range(matrix.shape[1]))
        ax6.set_xticklabels(xtick_labels[: matrix.shape[1]], rotation=90, fontsize=6)
        sep = max(1, len(panel6))
        for i in range(1, len(mismatch_runs)):
            ax6.axvline(i * sep - 0.5, color="k", lw=0.6)
        for i, mr in enumerate(mismatch_runs):
            ax6.text(
                (i + 0.5) * sep - 0.5,
                -0.65,
                mr,
                transform=ax6.get_xaxis_transform(),
                ha="center",
                fontsize=7,
            )
        fig6.colorbar(im, ax=ax6, fraction=0.02, label="final err (m)")
        fig6.tight_layout()
        _paper_save(fig6, os.path.join(plot_dir, "fig_mismatch_heatmap"))

    # --- Figure 7: upside-down (nominal only) ---
    nom = "nominal"
    inv_names = [scenario_names_all[i] for i in inverted_idx if i < len(scenario_names_all)]
    if nom in all_ep_data and inv_names:
        n_inv = min(len(inv_names), 4)
        fig7, axes7 = plt.subplots(n_inv, 2, figsize=(3.6, 1.35 * n_inv + 0.8), squeeze=False)
        for ri, scen in enumerate(inv_names[:n_inv]):
            ed = all_ep_data[nom].get(scen, {})
            for ci, kind in enumerate(["z", "tilt"]):
                ax = axes7[ri, ci]
                tmax = 0.0
                for ctrl in controllers:
                    d = ed.get(ctrl)
                    ps = PAPER_CTRL_STYLE[ctrl]
                    if d is None:
                        continue
                    obs = d["obs"]
                    t = np.arange(len(obs)) * sim_dt
                    tmax = max(tmax, float(t[-1]) if len(t) else 0.0)
                    if kind == "z":
                        z = d["positions"][:, 2]
                        ax.plot(t, z, color=ps["color"], ls=ps["linestyle"], label=ctrl)
                    else:
                        tilt = cs._quat_to_tilt_deg(obs[:, 3], obs[:, 4], obs[:, 5], obs[:, 6])
                        ax.plot(t, tilt, color=ps["color"], ls=ps["linestyle"], label=ctrl)
                    row = _row(nom, scen)
                    st = int(row.get(f"{ctrl}_min_steps", max_steps)) if row else max_steps
                    if st < max_steps:
                        ax.text(
                            0.05,
                            0.95 - 0.08 * controllers.index(ctrl),
                            f"{ctrl}: CRASH at t={st * sim_dt:.1f}s",
                            transform=ax.transAxes,
                            fontsize=6,
                            color=ps["color"],
                        )
                ax.set_xlim(0, max(tmax, sim_dt))
                if kind == "z":
                    ax.set_ylim(0.0, PAPER_PLOT_METERS_YMAX)
                ax.set_title(_abbrev(scen, 18))
                if ci == 0:
                    ax.set_ylabel("z (m)" if kind == "z" else "tilt (deg)")
        h7 = [
            Line2D([0], [0], color=PAPER_CTRL_STYLE[c]["color"], ls=PAPER_CTRL_STYLE[c]["linestyle"], label=c)
            for c in controllers
        ]
        fig7.legend(handles=h7, loc="lower center", ncol=2, bbox_to_anchor=(0.5, -0.02))
        fig7.tight_layout(rect=[0, 0.05, 1, 1])
        _paper_save(fig7, os.path.join(plot_dir, "fig_upside_down"))

    print(f"Paper figures written under: {plot_dir}")


def write_comparison_summary_xlsx(
    out_path: str,
    all_summaries: List[dict],
    controllers: List[str],
    scfg: Dict[str, Any],
    template_path: Optional[str] = None,
) -> None:
    """Write MAE, wall time, and final error grids (controllers × comparison runs × scenarios).

    Layout matches ``template.xlsx``: one sheet per metric, rows = scenario names,
    column blocks per controller and sub-columns per ``comparison_runs`` entry.
    ``template_path`` is reserved for future style copying from ``template.xlsx``.
    """
    _ = template_path
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Alignment, Font
    except ImportError as e:
        raise ImportError("Excel export requires openpyxl (pip install openpyxl).") from e

    run_order = [r["name"] for r in scfg.get("comparison_runs", [])]
    if not run_order:
        run_order = sorted({row["run"] for row in all_summaries})
    scenarios_cfg = scfg.get("scenarios", [])
    scenario_names = [s.get("name", f"S{i}") for i, s in enumerate(scenarios_cfg)]

    ctrl_order = ["RL", "RL+CBF", "MPC", "MPC+Con"]
    metrics: List[Tuple[str, str]] = [
        ("MAE", "mae"),
        ("Wall Time", "wall"),
        ("Final Error", "err"),
    ]

    def _lookup(run: str, scen: str, ctrl: str, suffix: str):
        for row in all_summaries:
            if row["run"] == run and row["name"] == scen:
                v = row.get(f"{ctrl}_{suffix}", float("nan"))
                if isinstance(v, (int, float)) and np.isfinite(v):
                    return float(v)
                return v
        return None

    wb = Workbook()
    wb.remove(wb.active)
    n_runs = len(run_order)
    bold = Font(bold=True)

    for sheet_title, key_suffix in metrics:
        ws = wb.create_sheet(sheet_title)
        ws.column_dimensions["A"].width = 44
        ws.cell(1, 1, "")
        ws.cell(2, 1, "Scenarios")
        for bi, cname in enumerate(ctrl_order):
            c0 = 2 + bi * n_runs
            c1 = c0 + n_runs - 1
            ws.merge_cells(start_row=1, start_column=c0, end_row=1, end_column=c1)
            h = ws.cell(1, c0, cname)
            h.font = bold
            h.alignment = Alignment(horizontal="center")
            for j, rn in enumerate(run_order):
                sub = ws.cell(2, c0 + j, _abbrev(str(rn), 36))
                sub.alignment = Alignment(horizontal="center", wrap_text=True, vertical="center")
        for si, scen in enumerate(scenario_names):
            rr = 3 + si
            label = ws.cell(rr, 1, scen)
            label.alignment = Alignment(wrap_text=True, vertical="center")
            for bi, cname in enumerate(ctrl_order):
                for j, rn in enumerate(run_order):
                    col = 2 + bi * n_runs + j
                    if cname not in controllers:
                        ws.cell(rr, col, "")
                        continue
                    val = _lookup(rn, scen, cname, key_suffix)
                    if val is not None and isinstance(val, (int, float)) and np.isfinite(val):
                        ws.cell(rr, col, val)
                    else:
                        ws.cell(rr, col, "")

    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    wb.save(out_path)


def write_summary_csv(
    path: str,
    all_summaries: List[dict],
    controllers: List[str],
    n_seeds: int,
) -> None:
    os.makedirs(os.path.dirname(path) or ".", exist_ok=True)
    base_fields = [
        "run",
        "scenario",
        "controller",
        "final_err",
        "mae",
        "total_reward",
        "steps",
        "wall_s",
        "thrust_asym",
        "t_conv",
    ]
    extra = (
        [
            "final_err_std",
            "mae_std",
            "rew_std",
            "steps_std",
            "wall_std",
            "thrust_asym_std",
            "t_conv_std",
        ]
        if n_seeds > 1
        else []
    )
    fieldnames = base_fields + extra
    with open(path, "w", newline="") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for row in all_summaries:
            for c in controllers:
                rec = {
                    "run": row["run"],
                    "scenario": row["name"],
                    "controller": c,
                    "final_err": row.get(f"{c}_err", ""),
                    "mae": row.get(f"{c}_mae", ""),
                    "total_reward": row.get(f"{c}_rew", ""),
                    "steps": row.get(f"{c}_steps", ""),
                    "wall_s": row.get(f"{c}_wall", ""),
                    "thrust_asym": row.get(f"{c}_thrust_asym", ""),
                    "t_conv": row.get(f"{c}_t_conv", ""),
                }
                if n_seeds > 1:
                    rec["final_err_std"] = row.get(f"{c}_err_std", "")
                    rec["mae_std"] = row.get(f"{c}_mae_std", "")
                    rec["rew_std"] = row.get(f"{c}_rew_std", "")
                    rec["steps_std"] = row.get(f"{c}_steps_std", "")
                    rec["wall_std"] = row.get(f"{c}_wall_std", "")
                    rec["thrust_asym_std"] = row.get(f"{c}_thrust_asym_std", "")
                    rec["t_conv_std"] = row.get(f"{c}_t_conv_std", "")
                w.writerow(rec)


def main() -> None:
    parser = argparse.ArgumentParser(
        description="Fair multi-run scenario comparison with split metric plots.",
    )
    parser.add_argument(
        "--scenarios_config",
        type=str,
        default=os.path.join(_REPO_ROOT, "configs", "scenarious.yaml"),
    )
    parser.add_argument("--checkpoint", type=str, default=None)
    parser.add_argument("--cbf_config", type=str, default=None)
    parser.add_argument("--mpc_config", type=str, default=None)
    parser.add_argument(
        "--controllers",
        nargs="+",
        default=None,
        choices=ALL_CONTROLLERS,
        help="Controllers to run (default: all)",
    )
    parser.add_argument(
        "--skip_controllers",
        nargs="+",
        default=None,
        choices=ALL_CONTROLLERS,
        help="Exclude these controllers (applied after --controllers).",
    )
    parser.add_argument("--n_seeds", type=int, default=1, help="Stochastic rollouts per scenario (mean±std).")
    parser.add_argument("--save_plots", action="store_true", default=False)
    parser.add_argument("--plot_dir", type=str, default=None)
    parser.add_argument("--comparison_base_seed", type=int, default=None)
    parser.add_argument(
        "--no_sync_stochastic_seeds",
        action="store_true",
        help="Do not reseed between controllers (legacy).",
    )
    parser.add_argument(
        "--paper_plots",
        action="store_true",
        help="Write publication figures to {plot_dir}/paper/ after evaluation.",
    )
    parser.add_argument(
        "--no_recompile_cbf",
        action="store_true",
        help="Reuse existing acados slack CBF .so if present (faster iterative runs).",
    )
    parser.add_argument(
        "--no-save-rollouts",
        action="store_true",
        help="Do not write rollout_data.pkl under each scenario folder (default: save).",
    )
    args = parser.parse_args()

    scfg = cs._load_scenarios_config(args.scenarios_config)
    runs = scfg.get("comparison_runs")
    if not runs:
        runs = [
            {
                "name": "nominal",
                "description": "implicit single run",
                "controller_model_overrides": {},
            }
        ]

    try:
        controllers = _resolve_controllers(args)
    except ValueError as e:
        parser.error(str(e))

    needs_rl = any(c in controllers for c in ["RL", "RL+CBF"])
    checkpoint = args.checkpoint or scfg.get("rl_policy", {}).get("checkpoint")
    if needs_rl and checkpoint is None:
        parser.error("--checkpoint required for RL controllers")

    plot_cfg = scfg.get("plotting", {})
    if args.plot_dir is None:
        args.plot_dir = plot_cfg.get("plot_dir", "fair_run_plots")
    if not args.save_plots and plot_cfg.get("save_plots", False):
        args.save_plots = True

    temp_files: List[str] = []
    all_summaries: List[dict] = []
    all_ep_data: Dict[str, Dict[str, Dict[str, dict]]] = {}
    try:
        for ri, run_cfg in enumerate(runs):
            rows, _, ep_map = _run_one_comparison_run(ri, run_cfg, scfg, args, temp_files)
            all_summaries.extend(rows)
            run_name = run_cfg.get("name", f"run_{ri}")
            all_ep_data[run_name] = ep_map
    finally:
        for p in temp_files:
            try:
                os.unlink(p)
            except OSError:
                pass

    n_seeds = max(1, int(args.n_seeds))
    col_w = 22
    W = 36 + col_w * len(controllers)
    print(f"\n{'='*W}")
    print("COMBINED SUMMARY (all comparison runs)")
    print(f"{'='*W}")
    header = f"{'Run':<18}{'Scenario':<30}" + "".join(f"{c:>{col_w}}" for c in controllers)
    print(header)
    print("-" * W)
    for row in all_summaries:
        line = f"{row['run']:<18}{row['name']:<30}"
        for c in controllers:
            err = row.get(f"{c}_err", float("nan"))
            steps = row.get(f"{c}_steps", 0)
            if n_seeds > 1 and f"{c}_err_std" in row:
                es = row[f"{c}_err_std"]
                line += f"{err:>6.3f}±{es:.2f}m {steps:>5.0f}st  "
            else:
                line += f"{err:>7.3f}m {steps:>5.0f}st  "
        print(line)
    print(f"{'='*W}")

    csv_path = os.path.join(args.plot_dir, "summary.csv")
    write_summary_csv(csv_path, all_summaries, controllers, n_seeds)
    print(f"Saved summary CSV to: {csv_path}")

    xlsx_path = os.path.join(args.plot_dir, "comparison_summary.xlsx")
    try:
        write_comparison_summary_xlsx(
            xlsx_path,
            all_summaries,
            controllers,
            scfg,
            os.path.join(_REPO_ROOT, "template.xlsx"),
        )
        print(f"Saved comparison Excel to: {xlsx_path}")
    except ImportError as e:
        print(f"Skipping Excel export: {e}")
    except OSError as e:
        print(f"Excel export failed: {e}")

    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_path = os.path.join(args.plot_dir, f"compare_fair_runs_log_{ts}.txt")
    os.makedirs(os.path.dirname(log_path) or ".", exist_ok=True)
    with open(log_path, "w") as f:
        f.write("compare_fair log\n")
        f.write(f"timestamp: {datetime.now().isoformat(timespec='seconds')}\n")
        f.write(f"scenarios_config: {args.scenarios_config}\n")
        f.write(f"checkpoint: {args.checkpoint}\n")
        f.write(f"controllers: {controllers}\n")
        f.write(f"n_seeds: {n_seeds}\n\n")
        for row in all_summaries:
            f.write(f"{row['run']} / {row['name']}\n")
            for c in controllers:
                if n_seeds > 1 and f"{c}_err_std" in row:
                    f.write(
                        f"  {c}: err={row.get(f'{c}_err', float('nan')):.4f} ± {row[f'{c}_err_std']:.4f}, "
                        f"mae={row.get(f'{c}_mae', float('nan')):.4f} ± {row.get(f'{c}_mae_std', float('nan')):.4f}, "
                        f"rew={row.get(f'{c}_rew', float('nan')):.3f}, "
                        f"steps={row.get(f'{c}_steps', 0)}, "
                        f"wall={row.get(f'{c}_wall', float('nan')):.4f}s\n",
                    )
                else:
                    f.write(
                        f"  {c}: err={row.get(f'{c}_err', float('nan')):.4f}, "
                        f"mae={row.get(f'{c}_mae', float('nan')):.4f}, "
                        f"rew={row.get(f'{c}_rew', float('nan')):.3f}, "
                        f"steps={row.get(f'{c}_steps', 0)}, "
                        f"wall={row.get(f'{c}_wall', float('nan')):.4f}s\n",
                    )
            f.write("\n")
    print(f"Saved log to: {log_path}")

    if args.paper_plots:
        generate_paper_plots(all_summaries, all_ep_data, scfg, args)


if __name__ == "__main__":
    main()
